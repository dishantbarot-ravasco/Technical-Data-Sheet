"""
Management command: seed_qap_templates

Reads SAMPLE_QAP.xlsx and populates QAPTemplate / QAPSection / QAPItem rows
for the three active belt categories: General Purpose, Heat Resistant, FR ISO.

Usage:
    python run_django.py seed_qap_templates --file /path/to/SAMPLE_QAP.xlsx
    python run_django.py seed_qap_templates --file /path/to/SAMPLE_QAP.xlsx --replace
    python run_django.py seed_qap_templates --file /path/to/SAMPLE_QAP.xlsx --dry-run

Sheet → Template mapping (OR and FR CAN CSA skipped — factory data not
received yet, so those two categories stay unseeded until it arrives):
    'cover grades for indus brute' → GP     General Purpose
    'FR grade except CAN CSA'      → FR_ISO Fire Resistant (ISO)
    'HR cover grades'              → HR     Heat Resistant

Column layout (fixed, confirmed from actual file):
    0  SN                 5  Quantum M      10  D* (skip)
    1  Component          6  Quantum S/C    11  Agency M
    2  Characteristic     7  Reference docs 12  Agency S
    3  Class              8  Acceptance     13  Agency C
    4  Type of check      9  Format         14  Remarks
"""

import re
from django.core.management.base import BaseCommand, CommandError

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ─── Sheet → template config ──────────────────────────────────────────────────
SHEET_MAP = {
    'cover grades for indus brute': {'category': 'GP',     'display_name': 'General Purpose'},
    'FR grade except CAN CSA':      {'category': 'FR_ISO', 'display_name': 'Fire Resistant (ISO)'},
    'HR cover grades':              {'category': 'HR',     'display_name': 'Heat Resistant'},
}

# Rows to skip — these are page-break signature lines or legal notes
SKIP_PHRASES = ('customer signature', 'notes:', 'repair norm', 'belt must be offered')

# Section header: SN column starts with "X.0" (e.g. "1.0 RAW MATERIAL")
SECTION_RE = re.compile(r'^(\d+\.0)\b(.*)')
# Item row: SN is "X.Y" or "X.Ya" (e.g. "1.1", "1.11", "2.1a", "4", "5")
ITEM_RE    = re.compile(r'^\d+(\.\d+[a-z]?)?$')


def _val(row, col):
    """Return stripped string from row[col], or '' if missing / None."""
    if col >= len(row):
        return ''
    v = row[col].value
    return str(v).strip() if v is not None else ''


def _agency(row):
    """
    Combine the three Agency columns (M=11, S=12, C=13) into one string.

    '-' is a MEANINGFUL value here (the source document literally prints a
    dash for "not applicable"), not a blank marker like it is in the Quantum
    S/C column - so it must be kept, not dropped. Only a genuinely empty
    cell is skipped.
    """
    m = _val(row, 11)
    s = _val(row, 12)
    c = _val(row, 13)
    parts = []
    for label, val in (('M', m), ('S', s), ('C', c)):
        if val:
            parts.append(f'{label}:{val}')
    return ' / '.join(parts) if parts else ''


def _row_fields(row):
    """
    Every column of ONE physical spreadsheet row, exactly as it appears -
    blank cells stay '' (they mean "merged with the row above" once this
    becomes a QAPItemSubRow; see that model's docstring for why this matters).
    """
    return {
        'characteristic':    _val(row, 2),
        'check_class':       _val(row, 3),
        'type_of_check':     _val(row, 4),
        'quantum_m':         _val(row, 5),
        'quantum_sc':        _val(row, 6),
        'reference_docs':    _val(row, 7),
        'acceptance_norms':  _val(row, 8),
        'format_of_records': _val(row, 9),
        'record_mark':       _val(row, 10),   # 'D' column - '√' tick mark
        'agency':            _agency(row),
        'remarks':           _val(row, 14),
    }


def _parse_sheet(ws):
    """
    Parse one worksheet into a flat list of dicts:
        {'type': 'section', 'code': '1.0', 'name': 'RAW MATERIAL', 'sort': N}
        {'type': 'item', 'section_code': '1.0', 'sn': '1.1', 'component': ...,
         ...row 0's fields..., 'sub_rows': [ {...row 1's fields...}, {...row 2...}, ... ]}

    Sub-characteristic rows (empty SN, sub-letter checks like "b) Ash Content")
    become entries in the parent item's 'sub_rows' list, each keeping its OWN
    class/type/quantum/reference/acceptance/agency exactly as that row had it
    (blank where the row was blank) - these are NOT merged into one string
    anymore, since several items (e.g. "1.1 Raw Rubber", "3.5 Cover Rubber
    Properties") have sub-rows that genuinely differ from the first row in
    Type of Check, Quantum, Reference Documents, or Acceptance Norms.
    """
    all_rows = list(ws.iter_rows())

    # Find the header row (the one where col 0 value is "SN")
    data_start = None
    for idx, row in enumerate(all_rows):
        if _val(row, 0).upper() in ('SN', 'S.N', 'S.NO'):
            data_start = idx + 1   # data starts the row after column headers
            break
    if data_start is None:
        return []

    # Skip the sub-header row (row 6: 'M', 'S/C' for quantum) and
    # the column-number row (row 7: '1','2','3',...) — both come right after
    # the header row. We detect them by checking col 0 is empty or numeric only.
    while data_start < len(all_rows):
        v = _val(all_rows[data_start], 0)
        if not v or v.isdigit():
            data_start += 1
        else:
            break

    results       = []
    sort_counter  = 0
    current_sec   = None
    last_item_idx = None   # index into results[] of the last item added

    for row in all_rows[data_start:]:
        sn_raw = _val(row, 0)

        # ── Skip completely empty rows ───────────────────────────────────────
        if not sn_raw and not _val(row, 1) and not _val(row, 2):
            continue

        # ── Skip signature / notes rows ─────────────────────────────────────
        sn_lower = sn_raw.lower()
        if any(phrase in sn_lower for phrase in SKIP_PHRASES):
            continue
        comp_lower = _val(row, 1).lower()
        if len(sn_raw) > 80:   # long notes text in SN column
            continue

        # ── Section header: "1.0 RAW MATERIAL" ──────────────────────────────
        m = SECTION_RE.match(sn_raw)
        if m:
            code = m.group(1)               # "1.0"
            name = m.group(2).strip()        # "RAW MATERIAL"
            sort_counter += 1
            current_sec   = code
            last_item_idx = None
            results.append({
                'type': 'section',
                'code': code,
                'name': name,
                'sort': sort_counter,
            })
            continue

        # ── Item row: SN like "1.1", "1.11", "3.1a", "4", "5" ──────────────
        if ITEM_RE.match(sn_raw):
            # Auto-create a section if this item has no parent (items 4, 5, etc.)
            if current_sec is None:
                sort_counter += 1
                synthetic_code = f'{sn_raw}.0'
                current_sec    = synthetic_code
                results.append({
                    'type': 'section',
                    'code': synthetic_code,
                    'name': 'Other',
                    'sort': sort_counter,
                })

            sort_counter += 1
            is_static = (current_sec == '1.0')
            item = {
                'type':              'item',
                'section_code':      current_sec,
                'sn':                sn_raw,
                'component':         _val(row, 1),
                'is_static':         is_static,
                'sort':              sort_counter,
                'sub_rows':          [],
                **_row_fields(row),
            }
            last_item_idx = len(results)
            results.append(item)
            continue

        # ── Sub-characteristic row (empty SN, "b) Ash Content" style) ────────
        # Kept as its own row (with its own class/type/quantum/reference/
        # acceptance/agency, blank where the source cell was blank) instead of
        # merging just the characteristic text into the parent - see
        # QAPItemSubRow's docstring for why several items need this.
        if not sn_raw and _val(row, 2) and last_item_idx is not None:
            prev = results[last_item_idx]
            if prev['type'] == 'item':
                prev['sub_rows'].append(_row_fields(row))
            continue

        # Anything else — skip
    return results


class Command(BaseCommand):
    help = 'Seed QAP template data from SAMPLE_QAP.xlsx into the database.'

    def add_arguments(self, parser):
        parser.add_argument('--file',    required=True,       help='Path to SAMPLE_QAP.xlsx')
        parser.add_argument('--replace', action='store_true', help='Wipe and re-seed if templates already exist')
        parser.add_argument('--dry-run', action='store_true', help='Print what would be created without touching the DB')

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError('openpyxl is not installed. Run: pip install openpyxl')

        from apps.core.models import QAPTemplate, QAPSection, QAPItem, QAPItemSubRow

        file_path = options['file']
        replace   = options['replace']
        dry_run   = options['dry_run']

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'File not found: {file_path}')
        except Exception as e:
            raise CommandError(f'Could not open Excel file: {e}')

        self.stdout.write(f'Opened: {file_path}')
        self.stdout.write(f'Available sheets: {wb.sheetnames}')

        # NOTE: --replace no longer deletes the QAPTemplate rows themselves -
        # only wipes their sections (which cascades to items/sub-rows). This
        # keeps each category's template PK stable across a re-seed, so any
        # QAPRecord already pointing at it (generated_at/doc_number history
        # for a previously-downloaded QAP) doesn't get its FK nulled out by
        # QAPRecord.template's on_delete=SET_NULL.
        if replace and not dry_run:
            cats = [v['category'] for v in SHEET_MAP.values()]
            deleted = QAPSection.objects.filter(template__category__in=cats).delete()
            self.stdout.write(self.style.WARNING(f'Wiped existing QAP sections/items: {deleted}'))

        totals = {'templates': 0, 'sections': 0, 'items': 0, 'sub_rows': 0}

        for sheet_name, cfg in SHEET_MAP.items():
            category, display_name = cfg['category'], cfg['display_name']

            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(
                    f'  Sheet "{sheet_name}" not found — skipping {category}'))
                continue

            existing_template = QAPTemplate.objects.filter(category=category).first()
            if existing_template and existing_template.sections.exists() and not replace:
                self.stdout.write(
                    f'  {category}: already seeded — skipping (use --replace to overwrite)')
                continue

            self.stdout.write(f'\n  Parsing sheet: "{sheet_name}" -> {category}')
            ws   = wb[sheet_name]
            rows = _parse_sheet(ws)

            sections = [r for r in rows if r['type'] == 'section']
            items    = [r for r in rows if r['type'] == 'item']
            self.stdout.write(f'    Found {len(sections)} sections, {len(items)} items')

            if dry_run:
                for r in rows[:20]:
                    self.stdout.write(f'      {r}')
                if len(rows) > 20:
                    self.stdout.write(f'      ... ({len(rows) - 20} more rows)')
                continue

            template, created = QAPTemplate.objects.get_or_create(
                category=category,
                defaults={'display_name': display_name, 'is_active': True},
            )
            if created:
                totals['templates'] += 1

            section_objs = {}
            for sec in sections:
                obj = QAPSection.objects.create(
                    template=template,
                    section_code=sec['code'],
                    section_name=sec['name'],
                    sort_order=sec['sort'],
                )
                section_objs[sec['code']] = obj
                totals['sections'] += 1

            for item in items:
                sec_obj = section_objs.get(item['section_code'])
                if sec_obj is None:
                    self.stdout.write(self.style.WARNING(
                        f'    Orphan item {item["sn"]} (section {item["section_code"]}) — skipping'))
                    continue
                item_obj = QAPItem.objects.create(
                    section=sec_obj,
                    sn=item['sn'],
                    component=item['component'],
                    characteristic=item['characteristic'],
                    check_class=item['check_class'],
                    type_of_check=item['type_of_check'],
                    quantum_m=item['quantum_m'],
                    quantum_sc=item['quantum_sc'],
                    reference_docs=item['reference_docs'],
                    acceptance_norms=item['acceptance_norms'],
                    format_of_records=item['format_of_records'],
                    record_mark=item['record_mark'],
                    agency=item['agency'],
                    remarks=item['remarks'],
                    is_static=item['is_static'],
                    sort_order=item['sort'],
                )
                totals['items'] += 1

                for i, sub in enumerate(item['sub_rows'], start=1):
                    QAPItemSubRow.objects.create(
                        item=item_obj,
                        characteristic=sub['characteristic'],
                        check_class=sub['check_class'],
                        type_of_check=sub['type_of_check'],
                        quantum_m=sub['quantum_m'],
                        quantum_sc=sub['quantum_sc'],
                        reference_docs=sub['reference_docs'],
                        acceptance_norms=sub['acceptance_norms'],
                        format_of_records=sub['format_of_records'],
                        record_mark=sub['record_mark'],
                        agency=sub['agency'],
                        remarks=sub['remarks'],
                        sort_order=i,
                    )
                    totals['sub_rows'] += 1

            self.stdout.write(self.style.SUCCESS(
                f'    Created "{display_name}" — {len(sections)} sections, {len(items)} items, '
                f'{sum(len(it["sub_rows"]) for it in items)} sub-rows'))

        self.stdout.write('')
        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN complete — no data written.'))
        else:
            self.stdout.write(self.style.SUCCESS(
                f'Done. Created: {totals["templates"]} templates, '
                f'{totals["sections"]} sections, {totals["items"]} items, '
                f'{totals["sub_rows"]} sub-rows.'))
